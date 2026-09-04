"""
Generates the Group Performance Dashboard's 3-pillar Excel data-entry template,
pre-filled with 5 quarters of dummy data (Q2 FY2025 -> Q2 FY2026).

Anchor values for Q1 FY2026 are pulled straight out of supabase/seed.sql wherever that quarter
is already seeded there, so the template stays consistent with what's already live in the
dashboard; every other quarter/row is a smooth, deterministic, plausible projection around that
anchor (or a reasonable fabricated base, for the handful of metric keys the app's own README
documents as "entry-only, never seeded" -- revenue_by_source, expense_by_category,
admin/personnel expense detail, pl_detail, receivables_aging).

Also regenerates src/lib/templateFields.generated.ts from the exact same manifest, so the
workbook this script writes and the parser that reads it back (src/lib/excelTemplate.ts) can
never drift out of sync. Run from the repo root: `python3 scripts/generate_data_entry_template.py`
(needs `openpyxl` -- `pip install openpyxl`).
"""
import os
import re
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SEED_SQL_PATH = os.path.join(REPO_ROOT, "supabase", "seed.sql")
OUT_DIR = os.path.join(REPO_ROOT, "scripts", "output")
TS_MANIFEST_PATH = os.path.join(REPO_ROOT, "src", "lib", "templateFields.generated.ts")

QUARTERS = ["Q2FY25", "Q3FY25", "Q4FY25", "Q1FY26", "Q2FY26"]
QUARTER_LABELS = ["Q2 FY2025", "Q3 FY2025", "Q4 FY2025", "Q1 FY2026", "Q2 FY2026"]
ANCHOR_IDX = 3  # Q1FY26 is the 4th column - the one with real seed.sql data behind it

# ---------------------------------------------------------------------------
# 1. Parse Q1FY26 anchors straight out of supabase/seed.sql
# ---------------------------------------------------------------------------
anchors = {}  # (metric_key, dim, dim2) -> float
kpi_anchors = {}  # kpi_id -> (ytd_target, ytd_actual)

with open(SEED_SQL_PATH) as f:
    for line in f:
        if "'Q1FY26'," not in line:
            continue
        m = re.search(
            r"'Q1FY26', '([a-z_]+)', '([^']*)', '([^']*)', ([\-\d.]+|null)",
            line,
        )
        if m:
            key, dim, dim2, val = m.groups()
            if val != "null":
                anchors[(key, dim, dim2)] = float(val)
        m2 = re.search(r"'(KPI\d+)', ?'HQ','Q1FY26', ([\-\d.]+|null), ([\-\d.]+|null)", line)
        if m2:
            kid, tgt, act = m2.groups()
            kpi_anchors[kid] = (
                float(tgt) if tgt != "null" else None,
                float(act) if act != "null" else None,
            )

print(f"Parsed {len(anchors)} detail_metrics anchors, {len(kpi_anchors)} KPI anchors")


def _parse_sql_tuple(s):
    """Split a SQL VALUES (...) tuple's inner content on top-level commas, respecting
    single-quoted strings (with '' as the escaped quote)."""
    parts, cur, in_str, i = [], "", False, 0
    while i < len(s):
        c = s[i]
        if in_str:
            if c == "'" and (i + 1 >= len(s) or s[i + 1] != "'"):
                in_str = False
            elif c == "'" and s[i + 1] == "'":
                cur += "'"
                i += 1
            else:
                cur += c
        else:
            if c == "'":
                in_str = True
            elif c == ",":
                parts.append(cur.strip())
                cur = ""
                i += 1
                continue
            else:
                cur += c
        i += 1
    parts.append(cur.strip())
    return parts


def _sql_val(tok):
    tok = tok.strip()
    return None if tok.lower() == "null" else tok


# record_anchors: (record_type, category_or_empty, label) -> (value_num, value_num2, text_note),
# parsed the same way as the detail_metrics anchors above but for detail_records -- powers the
# managed_entity_kpi / governance_kpi / financial_breakdown sections below, all of which need
# more than a single (metric_key, dim, dim2) -> value lookup.
record_anchors = {}
with open(SEED_SQL_PATH) as f:
    for line in f:
        if "insert into detail_records" not in line or "'Q1FY26'" not in line:
            continue
        m = re.search(r"values \((.*)\);\s*$", line.strip())
        if not m:
            continue
        toks = _parse_sql_tuple(m.group(1))
        if len(toks) != 10:
            continue
        _id, _entity, period_id, record_type, label, category, value_num, value_num2, text_note, _flag = [_sql_val(t) for t in toks]
        if period_id != "Q1FY26":
            continue
        record_anchors[(record_type, category or "", label)] = (
            float(value_num) if value_num is not None else None,
            float(value_num2) if value_num2 is not None else None,
            text_note,
        )

print(f"Parsed {len(record_anchors)} detail_records anchors")

# managed_entity_kpi's per-(entity, item) catalog -- entity + item description together identify
# the row, matching supabase/seed.sql. New items are rare and still need adding here + in-app
# (their "no|section|fyTarget|ytdTarget|ytdActual" text stays entered directly in CP004) before
# Rating/Weighted for that item become uploadable through the template.
MANAGED_ENTITY_KPI_ITEMS = [
    ("SJPP", "Sessions with FIs, GoM agencies, & SME enablers (SMEs or business associations)"),
    ("SJPP", "Processing of applications until GNL issued"),
    ("SJPP", "Processing of issuance & renewal of guarantees request from the FIs"),
    ("SJPP", "Issuance of GC for DPGS customers"),
    ("SJPP", "Number of Malaysian businesses approved for SJPP-guaranteed financing"),
    ("SJKP", "Processing and issuance of NOG for applications"),
    ("SJKP", "Verification and reconciliation of monthly reports from participating FIs"),
    ("SJKP", "Engagements with FIs and key stakeholders through regular sessions, events, and collaborations"),
    ("SJKP", "Facilitate home financing access to segments identified in NB — assist 1st time house buyer to gain access to home financing under the SJKP schemes"),
    ("DanaInfra", "Fundraising / financing activities (applicable for new mandate)"),
    ("DanaInfra", "Disbursement of funds made on disbursement date (applicable for existing mandates)"),
    ("DanaInfra", "Obtain optimal financing cost in fundraising"),
    ("DanaHarta", "Finalisation of outstanding recovery / settlement cases within the period"),
    ("DanaHarta", "Submission of quarterly recovery and compliance reports to relevant authorities"),
    ("DanaHarta", "Engagement sessions with government agencies and financial institutions"),
]

# governance_kpi's 5 fixed components (20% weight each) -- same order as their "no" prefix in
# textNote, matching supabase/seed.sql and CP004's own rendering order.
GOVERNANCE_KPI_ITEMS = [
    "Audit closure against the total audit observations due for implementation within the period",
    "Achieve satisfactory rating (rating 3) in audit findings within the period",
    "Implementation of Risk Action Plan (\"RAP\") within the timeline",
    "Completion of Organisational Anti-Corruption Strategy (\"OACS\") initiatives as planned",
    "Zero cases under Malaysian Anti-Corruption Commission (\"MACC\")",
]

# related_party_txn's category + "subheading|party" catalog, matching supabase/seed.sql exactly
# (dimension2 packs "subheading|party" -- see details.tsx's relatedPartyTransactions comment).
RELATED_PARTY_TXN_ROWS = [
    ("A. Subsidiary companies", "Management fee|PAMSB", 10277),
    ("B. Related corporations", "Management fee|PDNB", 245),
    ("B. Related corporations", "Management fee|SJPP", 35705),
    ("B. Related corporations", "Management fee|SJKP", 3587),
    ("B. Related corporations", "Management fee|DINB", 3300),
    ("B. Related corporations", "Management fee|GOVCO", 125),
    ("B. Related corporations", "Advisory / outsourcing services|DINB", 7),
]

PBT_BREAKDOWN_ITEMS = ["Revenue", "Other Income", "Total Income", "Expenses", "Profit Before Tax"]
CIR_BREAKDOWN_ITEMS = [
    "Personnel expenses", "Administrative expenses", "Professional fees", "Depreciation",
    "Other expenses", "Total Cost", "Total Income", "Cost-to-Income Ratio",
]

KPI_NAMES = {
    "KPI1": "Profit Before Tax (PBT)", "KPI2": "Cost-to-Income Ratio",
    "KPI3": "Managed Entities Rating", "KPI4": "Governance Index",
    "KPI5": "External Client Satisfaction", "KPI6": "Time Charter Compliance",
    "KPI7": "Process Improvements", "KPI8": "New Technology Implementation",
    "KPI9": "Recruitment Efficiency Index", "KPI10": "People Development Programme",
    "KPI11": "Bumiputera Procurement", "KPI12": "Bumiputera Composition",
    "KPI13": "Bumiputera Training",
}
# KPI4/5/6/7/10/13 are legitimately not-measurable most quarters (annual/bi-annual/not-yet-due)
KPI_SKIP_QUARTERS = {
    "KPI4": {0, 1, 3, 4},          # governance: annual, assessed Q4 only
    "KPI5": {1, 3},                # bi-annual survey: Q2 and Q4 rounds only (idx 0,2,4 here)
    "KPI6": {0},                   # baseline framework still being set up in Q2FY25
    "KPI7": {0},                   # initiatives hadn't started reporting yet
    "KPI10": {0},                  # programmes commenced from Q3FY25
    "KPI13": {0, 3},               # training commences properly from Q2 each cycle
}


def series(anchor, direction, spread=0.09, decimals=1, floor=None):
    """5-point deterministic series around an anchor at ANCHOR_IDX, drifting per `direction`."""
    rnd = random.Random(round(anchor * 1000) + hash(direction) % 97)
    drift = {"growth": 0.045, "cost": 0.02, "flat": 0.0, "decline": -0.03}[direction]
    vals = [None] * 5
    vals[ANCHOR_IDX] = anchor
    cur = anchor
    for i in range(ANCHOR_IDX - 1, -1, -1):
        cur = cur / (1 + drift + rnd.uniform(-spread, spread))
        vals[i] = cur
    cur = anchor
    for i in range(ANCHOR_IDX + 1, 5):
        cur = cur * (1 + drift + rnd.uniform(-spread, spread))
        vals[i] = cur
    if floor is not None:
        vals = [max(floor, v) for v in vals]
    return [round(v, decimals) for v in vals]


def fresh_series(base, direction, spread=0.10, decimals=1, floor=0, seed_key=""):
    """5-point series with no real anchor at all -- used for the never-seeded metric keys."""
    rnd = random.Random(hash(seed_key) % (2**31))
    drift = {"growth": 0.05, "cost": 0.02, "flat": 0.0, "decline": -0.03}[direction]
    vals, cur = [], base
    for i in range(5):
        cur = cur * (1 + drift + rnd.uniform(-spread, spread)) if i else base
        vals.append(cur)
    return [round(max(floor, v), decimals) for v in vals]


# ---------------------------------------------------------------------------
# 2. Manifest: every field the dashboard actually reads, grouped into the app's own
#    3-pillar navigation (Corporate Performance / Financial Health / Resource & People)
# ---------------------------------------------------------------------------

def cp_rows():
    rows = []
    rows.append(("section", "KPI Scorecard", None))
    for i in range(1, 14):
        kid = f"KPI{i}"
        tgt, act = kpi_anchors.get(kid, (None, None))
        vals = [None] * 5
        skip = KPI_SKIP_QUARTERS.get(kid, set())
        if act is not None:
            base_series = series(act, "growth" if kid not in ("KPI2",) else "decline", decimals=1)
        else:
            base_series = fresh_series(50 if kid != "KPI13" else 100, "growth", seed_key=kid)
        for qi in range(5):
            vals[qi] = None if qi in skip else base_series[qi]
        rows.append(("kpi", kid, KPI_NAMES[kid], vals))

    # managed_entity_ratings (metric_key) and governance_index (record_type) were both retired
    # from the live schema in favour of managed_entity_kpi / governance_kpi below -- the roll-up
    # summary (met/not-met/achievement, or the governance total) is now derived entirely from
    # those per-item rows rather than entered as its own separate figure.
    rows.append(("section", "Managed Entities KPI Detail (KPI 3 support)", None))
    for entity, item in MANAGED_ENTITY_KPI_ITEMS:
        rating_a, weighted_a, _ = record_anchors.get(("managed_entity_kpi", entity, item), (None, None, None))
        rating_vals = series(rating_a, "flat", spread=0.1, decimals=0, floor=1) if rating_a is not None else fresh_series(3, "flat", decimals=0, floor=1, seed_key=entity + item + "rating")
        weighted_vals = series(weighted_a, "flat", spread=0.1, decimals=2, floor=0) if weighted_a is not None else fresh_series(0.2, "flat", decimals=2, floor=0, seed_key=entity + item + "weighted")
        display = f"{entity} — {item}"
        rows.append(("record", "managed_entity_kpi", item, entity, f"{display} — Rating", rating_vals, "valueNum"))
        rows.append(("record", "managed_entity_kpi", item, entity, f"{display} — Weighted", weighted_vals, "valueNum2"))

    # Only the weighted score (the number that actually drives the KPI4 roll-up) is templated --
    # fyTarget/ytdActual/achievement are free text per item (some are percentages, some are
    # "Rating 3.0", some are "0 cases", one goes "N/A -- No initiatives due" outside its window),
    # not a uniform numeric cell, so those stay entered directly in-app (CP004).
    rows.append(("section", "Governance KPI Detail (KPI 4 support)", None))
    for item in GOVERNANCE_KPI_ITEMS:
        _, weighted_a, _ = record_anchors.get(("governance_kpi", "", item), (None, None, None))
        vals = series(weighted_a, "flat", decimals=1, spread=0.08, floor=0) if weighted_a is not None else fresh_series(15, "flat", decimals=1, floor=0, seed_key=item)
        rows.append(("record", "governance_kpi", item, "", f"{item} — Weighted", vals, "valueNum2"))

    rows.append(("section", "Client Satisfaction (KPI 5 support)", None))
    cs_target, cs_actual, _ = record_anchors.get(("client_satisfaction", "", "External Client Satisfaction"), (None, None, None))
    for label_suffix, field, a, base in [("FY TARGET", "valueNum", cs_target, 4.7), ("YTD ACTUAL", "valueNum2", cs_actual, 4.5)]:
        vals = series(a, "flat", decimals=1, spread=0.03) if a is not None else fresh_series(base, "flat", decimals=1, seed_key="client_satisfaction" + field)
        rows.append(("record", "client_satisfaction", "External Client Satisfaction", "", f"External Client Satisfaction — {label_suffix}", vals, field))

    rows.append(("section", "Time Charter Compliance (KPI 6 support, % per department)", None))
    for dept in [
        "Administration & Security", "Credit", "Finance", "Guarantee Schemes - Claims", "Human Resources",
        "Information Technology", "Legal Affairs", "Operational Excellence", "Secretarial Services",
        "Syarikat Jaminan Kredit Perumahan (SJKP)", "Syarikat Jaminan Pembiayaan Perniagaan (SJPP)",
    ]:
        a = anchors.get(("time_charter_dept_score", dept, ""))
        vals = series(a, "flat", decimals=1, spread=0.02, floor=0) if a is not None else fresh_series(97, "flat", decimals=1, floor=0, seed_key=dept)
        rows.append(("metric", "time_charter_dept_score", dept, "", dept, vals))

    # Only weight/weighted -- the two numeric fields this template can carry (see excelTemplate.ts:
    # a metric row's uploaded value is a single number, no text). score/computation stay text-only,
    # entered in-app -- RecruitmentIndexScorecard.tsx already computes a sensible score % and
    # "score x weight%" fallback whenever those two text fields are blank, so this alone is enough
    # to stop CP007/RP001 showing "Component breakdown not tracked for this reporting period."
    rows.append(("section", "Recruitment Efficiency Index (KPI 9 support)", None))
    for name in ["Time to Hire (TTH)", "MRF Fulfilment Rate", "Quality of Hire", "Offer Acceptance Rate"]:
        for sub, dec in [("weight", 2), ("weighted", 4)]:
            a = anchors.get(("recruitment_index", name, sub))
            vals = series(a, "flat", decimals=dec, spread=0.08, floor=0) if a is not None else fresh_series(0.2, "flat", decimals=dec, floor=0, seed_key=name + sub)
            rows.append(("metric", "recruitment_index", name, sub, f"{name} — {sub.title()}", vals))

    rows.append(("section", "Bumiputera Procurement (KPI 11 support, RM mil)", None))
    for dept in ["Administration & Security", "Corporate Communications", "Information Technology"]:
        for sub, dirn in [("fy_target", "flat"), ("ytd_actual", "growth")]:
            a = anchors.get(("bumiputera_procurement", dept, sub))
            vals = series(a, dirn, decimals=2, spread=0.1) if a is not None else fresh_series(0.3, dirn, decimals=2, seed_key=dept + sub)
            rows.append(("metric", "bumiputera_procurement", dept, sub, f"{dept} — {sub.replace('_', ' ').upper()}", vals))

    rows.append(("section", "Bumiputera Training (KPI 13 support)", None))
    for sub, base in [("pool_identified", 145), ("attended_one", 30), ("attended_two_plus", 2)]:
        a = anchors.get(("bumiputera_training", sub, ""))
        vals = series(a, "growth", decimals=0, floor=0) if a is not None else fresh_series(base, "growth", decimals=0, floor=0, seed_key=sub)
        rows.append(("metric", "bumiputera_training", sub, "", sub.replace("_", " ").title(), vals))
    return rows


def fh_rows():
    rows = []
    rows.append(("section", "Financial Trend (RM mil / %)", None))
    for dim, dec in [("revenue", 1), ("pbt", 1), ("cir", 1), ("net_margin", 1)]:
        a = anchors.get(("financial_trend", dim, ""))
        dirn = "cost" if dim == "cir" else "growth"
        vals = series(a, dirn, decimals=dec) if a is not None else fresh_series(20, dirn, decimals=dec, seed_key=dim)
        rows.append(("metric", "financial_trend", dim, "", dim.replace("_", " ").upper(), vals))

    rows.append(("section", "Actual vs Budget vs Prior Year (RM mil)", None))
    for line in ["Revenue / Total Income", "Staff Cost", "Admin & Operating Cost", "Total Cost", "Profit Before Tax", "Net Profit Margin (%)"]:
        for sub, dirn in [("actual", "growth"), ("budget", "flat"), ("py", "flat")]:
            a = anchors.get(("actual_vs_budget", line, sub))
            vals = series(a, dirn, decimals=1, spread=0.06) if a is not None else fresh_series(10, dirn, decimals=1, seed_key=line + sub)
            rows.append(("metric", "actual_vs_budget", line, sub, f"{line} — {sub.upper()}", vals))

    rows.append(("section", "Revenue by Source (RM mil)", None))
    for label, dim, base in [
        ("Management fee — DanaHarta (investment activities)", "danaharta_mgmt_fee", 31.9),
        ("Management fee — GovCo", "govco_mgmt_fee", 0.1),
        ("Management fee — SJKP", "sjkp_mgmt_fee", 3.1),
        ("Management fee — SJPP", "sjpp_mgmt_fee", 31.9),
        ("Management fee — DanaInfra", "danainfra_mgmt_fee", 3.0),
        ("Fee from SAP services", "sap_services_fee", 2.4),
        ("Fee from outsourcing services", "outsourcing_services_fee", 1.8),
        ("Fee from secretarial services", "secretarial_services_fee", 0.6),
        ("Fee from corporate advisory services", "corporate_advisory_fee", 0.9),
        ("Fee from credit advisory services", "credit_advisory_fee", 0.7),
        ("Income from acquired loans (PAM)", "acquired_loans_income", 4.3),
    ]:
        # Actual and budget are two independent numbers per source (PFH002's Actual vs Budget
        # drill-down needs both) -- dimension2 carries which one, matching the live schema.
        for dim2, dirn in [("actual", "growth"), ("budget", "flat")]:
            a = anchors.get(("revenue_by_source", dim, dim2))
            vals = series(a, dirn, decimals=2, spread=0.08) if a is not None else fresh_series(base, dirn, decimals=2, seed_key=dim + dim2)
            rows.append(("metric", "revenue_by_source", dim, dim2, f"{label} — {dim2.upper()}", vals))

    rows.append(("section", "Expense by Category (RM mil)", None))
    for label, dim, base in [
        ("Administrative expenses", "admin_expenses", 2.3),
        ("Personnel expenses", "personnel_expenses", 16.8),
        ("Professional fees", "professional_fees", 0.4),
        ("Depreciation", "depreciation", 1.2),
        ("Depreciation of ROU asset", "depreciation_rou", 0.2),
        ("Other expenses", "other_expenses", 0.7),
        ("Interest expense (lease liability)", "interest_expense_lease", 0.1),
        ("Provision for impairment loss on receivables", "impairment_receivables", 0.2),
    ]:
        for dim2, dirn in [("actual", "cost"), ("budget", "flat")]:
            a = anchors.get(("expense_by_category", dim, dim2))
            vals = series(a, dirn, decimals=2, spread=0.08) if a is not None else fresh_series(base, dirn, decimals=2, seed_key=dim + dim2)
            rows.append(("metric", "expense_by_category", dim, dim2, f"{label} — {dim2.upper()}", vals))

    rows.append(("section", "Administrative Expense Detail (RM '000)", None))
    admin_items = [
        "Advertisement, printing and stationery", "Computer expenses", "Electricity", "Insurance",
        "Motor vehicle expenses", "Newspaper/periodicals", "Office maintenance and repairs",
        "Lease (service tax)", "Office equipment rental and maintenance", "Stamps and postages",
        "Telephone", "Travelling", "Entertainment", "Corporate Communication", "Service tax",
    ]
    for label in admin_items:
        vals = fresh_series(round(random.Random(label).uniform(8, 60), 0), "cost", decimals=0, seed_key=label)
        rows.append(("metric", "admin_expense_detail", label, "", label, vals))

    rows.append(("section", "Personnel Expense Detail (RM '000)", None))
    personnel_items = [
        "Salaries and wages", "Bonus provision", "EPF contribution", "SOCSO", "Staff training",
        "Staff welfare", "Medical expenses", "HRDF fee", "Director's benefits",
        "Director's meeting allowance", "MoF staff related cost",
    ]
    for label in personnel_items:
        base = 8500 if label == "Salaries and wages" else round(random.Random(label).uniform(20, 900), 0)
        vals = fresh_series(base, "cost", decimals=0, seed_key=label)
        rows.append(("metric", "personnel_expense_detail", label, "", label, vals))

    rows.append(("section", "P&L Detail below PBT (RM mil)", None))
    for label, dim, base, dirn in [
        ("Finance income", "finance_income", 1.8, "growth"),
        ("Other income", "other_income", 6.2, "growth"),
        ("Taxation", "taxation", -7.9, "cost"),
        ("Profit after tax", "profit_after_tax", 23.7, "growth"),
        ("Dividend", "dividend", 15.0, "flat"),
    ]:
        for dim2 in ("actual", "budget"):
            a = anchors.get(("pl_detail", dim, dim2))
            vals = series(a, dirn, decimals=1, spread=0.08, floor=-999) if a is not None else fresh_series(base, dirn, decimals=1, floor=-999, seed_key=dim + dim2)
            rows.append(("metric", "pl_detail", dim, dim2, f"{label} — {dim2.upper()}", vals))

    rows.append(("section", "Balance Sheet Trend (RM mil)", None))
    for dim, label in [("shareholders_fund", "Shareholders' Fund"), ("total_liabilities", "Total Liabilities")]:
        a = anchors.get(("balance_sheet", dim, ""))
        vals = series(a, "growth", decimals=1) if a is not None else fresh_series(500, "growth", decimals=1, seed_key=dim)
        rows.append(("metric", "balance_sheet", dim, "", label, vals))

    rows.append(("section", "Balance Sheet Line Items (RM mil)", None))
    for label, dim2 in [
        ("Cash & Equivalents", "asset"), ("Fixed Income Investments", "asset"),
        ("Receivables & Others", "asset"), ("Fixed Assets (Net)", "asset"),
        ("Trade & Other Payables", "liability"), ("Deferred Revenue", "liability"),
        ("Other Liabilities", "liability"),
    ]:
        a = anchors.get(("balance_sheet_lines", label, dim2))
        dirn = "growth" if dim2 == "asset" else "flat"
        vals = series(a, dirn, decimals=1) if a is not None else fresh_series(10, dirn, decimals=1, seed_key=label)
        rows.append(("metric", "balance_sheet_lines", label, dim2, f"{label} ({dim2})", vals))

    rows.append(("section", "Receivables Aging (RM '000)", None))
    for label, dim, base in [
        ("Current", "current", 420), ("1–30 days", "days_1_30", 180), ("31–60 days", "days_31_60", 95),
        ("61–90 days", "days_61_90", 40), ("91–120 days", "days_91_120", 15),
        ("Over 120 days (impaired)", "days_over_120", 8),
    ]:
        vals = fresh_series(base, "flat", decimals=0, seed_key=dim)
        rows.append(("metric", "receivables_aging", dim, "", label, vals))

    # Real disclosed figures, not a projectable trend -- only the two quarters actually filed
    # (Q4FY25, Q1FY26) get a value; every other column stays blank rather than fabricating a
    # related-party transaction that never happened.
    rows.append(("section", "Related Party Transactions (RM '000)", None))
    for category, subheading_party, q4fy25_val in RELATED_PARTY_TXN_ROWS:
        subheading, party = subheading_party.split("|")
        q1fy26_val = anchors.get(("related_party_txn", category, subheading_party))
        vals = [None, None, q4fy25_val, q1fy26_val, None]
        rows.append(("metric", "related_party_txn", category, subheading_party, f"{subheading} — {party} ({category})", vals))

    rows.append(("section", "PBT Breakdown (RM mil)", None))
    for item in PBT_BREAKDOWN_ITEMS:
        vn, vn2, note = record_anchors.get(("financial_breakdown", "PBT", item), (None, None, None))
        ytd_target = float(note) if note not in (None, "") else None
        fy_vals = series(vn, "flat", decimals=1, spread=0.05) if vn is not None else fresh_series(100, "flat", decimals=1, floor=-999, seed_key=item + "fy")
        ytdt_vals = series(ytd_target, "flat", decimals=1, spread=0.05) if ytd_target is not None else fresh_series(90, "flat", decimals=1, floor=-999, seed_key=item + "ytdt")
        ytda_vals = series(vn2, "growth", decimals=1, spread=0.08, floor=-999) if vn2 is not None else fresh_series(80, "growth", decimals=1, floor=-999, seed_key=item + "ytda")
        rows.append(("record", "financial_breakdown", item, "PBT", f"{item} — FY Target", fy_vals, "valueNum"))
        rows.append(("record", "financial_breakdown", item, "PBT", f"{item} — YTD Target", ytdt_vals, "textNote"))
        rows.append(("record", "financial_breakdown", item, "PBT", f"{item} — YTD Actual", ytda_vals, "valueNum2"))

    # CIR's rows only ever showed an actual-figures column (see details.tsx) -- no FY/YTD target.
    rows.append(("section", "CIR Breakdown (RM mil)", None))
    for item in CIR_BREAKDOWN_ITEMS:
        _, vn2, _ = record_anchors.get(("financial_breakdown", "CIR", item), (None, None, None))
        vals = series(vn2, "flat", decimals=1, spread=0.05, floor=-999) if vn2 is not None else fresh_series(15, "flat", decimals=1, floor=-999, seed_key=item + "cir")
        rows.append(("record", "financial_breakdown", item, "CIR", f"{item} — YTD Actual (CIR)", vals, "valueNum2"))
    return rows


def rp_rows():
    rows = []
    rows.append(("section", "Headcount Summary", None))
    for dim, label, dec in [
        ("total_employees", "Total Employees", 0), ("bumiputera", "Bumiputera", 0),
        ("non_bumiputera", "Non-Bumiputera", 0), ("approved_headcount", "Approved Headcount", 0),
        ("filled_position", "Filled Position", 0),
    ]:
        a = anchors.get(("headcount_summary", dim, ""))
        vals = series(a, "growth", decimals=dec, spread=0.03) if a is not None else fresh_series(200, "growth", decimals=dec, seed_key=dim)
        rows.append(("metric", "headcount_summary", dim, "", label, vals))

    rows.append(("section", "Gender Breakdown", None))
    for dim in ["male", "female"]:
        a = anchors.get(("gender_breakdown", dim, ""))
        vals = series(a, "flat", decimals=0, spread=0.04) if a is not None else fresh_series(100, "flat", decimals=0, seed_key=dim)
        rows.append(("metric", "gender_breakdown", dim, "", dim.title(), vals))

    rows.append(("section", "Average Age", None))
    a = anchors.get(("average_age", "avg", ""))
    vals = series(a, "flat", decimals=1, spread=0.01) if a is not None else fresh_series(38, "flat", decimals=1, seed_key="average_age")
    rows.append(("metric", "average_age", "avg", "", "Average Age (years)", vals))

    rows.append(("section", "Grade Breakdown (5 approved bands)", None))
    for dim in ["Top Management", "Senior Management", "Management", "Executive", "Non-Executive"]:
        a = anchors.get(("grade_breakdown", dim, ""))
        vals = series(a, "flat", decimals=0, spread=0.05) if a is not None else fresh_series(30, "flat", decimals=0, seed_key=dim)
        rows.append(("metric", "grade_breakdown", dim, "", dim, vals))

    rows.append(("section", "Age Breakdown (4 bands)", None))
    for dim in ["≤30", "31–40", "41–50", "51+"]:
        a = anchors.get(("age_breakdown", dim, ""))
        vals = series(a, "flat", decimals=0, spread=0.04) if a is not None else fresh_series(50, "flat", decimals=0, seed_key=dim)
        rows.append(("metric", "age_breakdown", dim, "", dim, vals))

    rows.append(("section", "Age × Gender Breakdown", None))
    for dim in ["≤30", "31–40", "41–50", "51+"]:
        for dim2 in ["male", "female"]:
            a = anchors.get(("age_gender_breakdown", dim, dim2))
            vals = series(a, "flat", decimals=0, spread=0.05) if a is not None else fresh_series(25, "flat", decimals=0, seed_key=dim + dim2)
            rows.append(("metric", "age_gender_breakdown", dim, dim2, f"{dim} — {dim2.title()}", vals))

    rows.append(("section", "Department Headcount", None))
    for dept in ["Finance", "Human Resource", "Corporate Performance", "IT & Digital", "Risk & Compliance"]:
        for dim2, label in [("approved", "Approved"), ("filled", "Filled")]:
            a = anchors.get(("dept_headcount", dept, dim2))
            vals = series(a, "flat", decimals=0, spread=0.04) if a is not None else fresh_series(18, "flat", decimals=0, seed_key=dept + dim2)
            rows.append(("metric", "dept_headcount", dept, dim2, f"{dept} — {label}", vals))

    rows.append(("section", "Resignations", None))
    a = anchors.get(("resigned", "count", ""))
    vals = series(a, "flat", decimals=0, spread=0.2, floor=0) if a is not None else fresh_series(4, "flat", decimals=0, floor=0, seed_key="resigned")
    rows.append(("metric", "resigned", "count", "", "Staff Resigned This Quarter", vals))
    return rows


SHEETS = {
    "Corporate Performance": cp_rows(),
    "Financial Health": fh_rows(),
    "Resource & People": rp_rows(),
}

for name, rows in SHEETS.items():
    n_data = sum(1 for r in rows if r[0] != "section")
    print(f"{name}: {n_data} data rows across {sum(1 for r in rows if r[0] == 'section')} sections")

# ---------------------------------------------------------------------------
# 3. Write one workbook per quarter — a separate file per reporting period,
#    each carrying just that one quarter's column, matching how a real
#    submission cycle works (one file per period, not one giant multi-quarter
#    file). The 3-pillar sheet layout and row manifest are identical across
#    every file; only which quarter column is present differs.
# ---------------------------------------------------------------------------
NAVY = "16324A"
ACCENT = "0E8F5C"
LIGHT = "EAF3EF"
BORDER = Border(bottom=Side(style="thin", color="D9E2E8"))

# Where each section's figures actually render — filled preparer's own reference, not parsed.
SCREEN_FOR = {
    "KPI Scorecard": "Main · every CP/PFH/RP screen for that KPI's own perspective",
    "Managed Entities KPI Detail (KPI 3 support)": "CP004 Mandate & Governance — Managed Entities Performance Summary table (per-entity, per-item Rating/Weighted; item catalog + targets stay in-app)",
    "Governance KPI Detail (KPI 4 support)": "CP004 Mandate & Governance — Governance Index panel (Weighted score only; FY Target/YTD Actual/Achievement text stays in-app)",
    "Client Satisfaction (KPI 5 support)": "CP005 Customer — External Client Satisfaction",
    "Time Charter Compliance (KPI 6 support, % per department)": "CP005 Customer — Time Charter Compliance department scoring + drilldown",
    "Recruitment Efficiency Index (KPI 9 support)": "CP007 Organisational Capacity · RP001 Section B — component scorecard",
    "Bumiputera Procurement (KPI 11 support, RM mil)": "CP008 Bumiputera Empowerment — Procurement tab",
    "Bumiputera Training (KPI 13 support)": "CP008 Bumiputera Empowerment — Training tab",
    "Financial Trend (RM mil / %)": "CP003 Financial Perspective (PBT/CIR charts) · PFH002 Financial Results QoQ",
    "Actual vs Budget vs Prior Year (RM mil)": "PFH003 Actual vs Budget vs PY",
    "Revenue by Source (RM mil)": "PFH002 Financial Results — Revenue drill-down (Actual vs Budget)",
    "Expense by Category (RM mil)": "PFH002 Financial Results — Expenses drill-down (Actual vs Budget)",
    "Administrative Expense Detail (RM '000)": "Not yet displayed — entered for a future MEC-report drill-down (see supabase/README.md §Entry-only)",
    "Personnel Expense Detail (RM '000)": "Not yet displayed — entered for a future MEC-report drill-down (see supabase/README.md §Entry-only)",
    "P&L Detail below PBT (RM mil)": "PFH002 Financial Results — income statement below PBT (Actual vs Budget)",
    "Balance Sheet Trend (RM mil)": "PFH004 Assets & Liabilities — balance sheet trend chart",
    "Balance Sheet Line Items (RM mil)": "PFH004 Assets & Liabilities — asset/liability breakdown",
    "Receivables Aging (RM '000)": "Not yet displayed — entered for a future MEC-report drill-down (see supabase/README.md §Entry-only)",
    "Related Party Transactions (RM '000)": "PFH005 Related Party Transactions table (only quarters actually filed carry a figure — leave others blank)",
    "PBT Breakdown (RM mil)": "CP003 Financial Perspective — PBT card drill-down",
    "CIR Breakdown (RM mil)": "CP003 Financial Perspective — CIR card drill-down",
    "Headcount Summary": "RP001 Total Headcount · RP002 Approved Headcount · CP008 Composition tab · Main pillar snapshot",
    "Average Age": "RP001A Staff Demographics — Age Profile Summary donut",
    "Gender Breakdown": "RP001A Staff Demographics — Section A",
    "Grade Breakdown (5 approved bands)": "RP001A Staff Demographics — Section B",
    "Age Breakdown (4 bands)": "RP001A Staff Demographics — Section C age heatmap",
    "Age × Gender Breakdown": "RP001A Staff Demographics — Section C grouped bar chart",
    "Department Headcount": "RP002 Approved Headcount & KPI 10 — Section A",
    "Resignations": "RP003 / RP004 Turnover Rate trend",
}


PILLAR_SLUG = {
    "Corporate Performance": "corporate_performance",
    "Financial Health": "financial_health",
    "Resource & People": "resource_people",
}

UNLOCKED = Protection(locked=False)
LOCKED = Protection(locked=True)


def build_workbook(qi: int, pillar: str) -> openpyxl.Workbook:
    """A single-quarter, single-pillar workbook — one sheet for `pillar`, one value column
    (QUARTER_LABELS[qi]). Split by pillar (rather than one workbook with 3 sheets) so each
    department's preparer only ever opens their own file. Every cell except the value column
    (and the free-text note column, for the KPI Scorecard section) is protected/locked, so a
    preparer can't rename a row, reorder a sheet, or fat-finger a label — Excel will refuse the
    edit unless they explicitly unprotect the sheet first."""
    quarter_label = QUARTER_LABELS[qi]
    rows = SHEETS[pillar]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    instructions = wb.create_sheet("Instructions")
    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions["A"].width = 100
    lines = [
        (f"{pillar} — Data Entry Template · {quarter_label}", 16, True, NAVY),
        ("", 11, False, "000000"),
        (f"One sheet, scoped to this dashboard pillar only: {pillar}.", 11, False, "333333"),
        (f"Each row is one figure the dashboard displays, for this one reporting period: {quarter_label}.", 11, False, "333333"),
        ("", 11, False, "000000"),
        ("HOW TO USE", 12, True, NAVY),
        ("1. Fill in or correct the shaded value column — every other cell is protected and can't be edited.", 11, False, "333333"),
        ("2. Leave a cell blank if that figure genuinely wasn't measured this quarter (e.g. an annual KPI outside Q4).", 11, False, "333333"),
        ("3. Upload this file from Data Entry — the quarter it writes to is read from the column header, not a dropdown.", 11, False, "333333"),
        ("4. The KPI Scorecard section (Corporate Performance sheet only) routes through the checker queue; every other row saves directly.", 11, False, "333333"),
        ("", 11, False, "000000"),
        ("SHEET IS PROTECTED", 12, True, NAVY),
        ("Row labels, sub-categories, headers and the reference remark column are locked so they can't be renamed or", 11, False, "333333"),
        ("reordered by accident. Only the value column (and the note column on KPI rows) accepts edits. Preparers who", 11, False, "333333"),
        ("genuinely need to change the sheet layout can turn this off from Excel's Review → Unprotect Sheet menu — no", 11, False, "333333"),
        ("password is set.", 11, False, "333333"),
        ("", 11, False, "000000"),
        ("WHAT'S PRE-FILLED", 12, True, NAVY),
        (f"This copy already carries dummy data for {quarter_label} so it's ready to test-upload as-is.", 11, False, "333333"),
        ("Q1 FY2026's copy matches what's already seeded in the live dashboard; every other quarter is a plausible projection.", 11, False, "333333"),
        ("Replace with real figures before using this for an actual reporting cycle.", 11, False, "333333"),
        ("", 11, False, "000000"),
        ("NOT COVERED HERE", 12, True, NAVY),
        ("Initiative lists (Process/Tech Initiatives, People Development Programme) stay entered directly in-app — each row", 11, False, "333333"),
        ("is a name, a start/end date and a status, not one number per quarter.", 11, False, "333333"),
        ("For Managed Entities KPI Detail and Governance KPI Detail, only the scored figures (Rating/Weighted) are here —", 11, False, "333333"),
        ("the item catalog (No./Section/description) and its FY/YTD Target and YTD Actual wording (a mix of %, ratings and", 11, False, "333333"),
        ("counts, not a uniform number) stay entered directly in-app on CP004.", 11, False, "333333"),
        ("The Variance Commentary notes (PFH003) are also entered directly in-app — a commentary sentence per line, not a", 11, False, "333333"),
        ("number, so it doesn't fit this template's one-number-per-quarter column.", 11, False, "333333"),
        ("No conditional formatting or colour-coded cells are used anywhere in this workbook — every figure is a plain", 11, False, "333333"),
        ("number, entered here and read as-is; status colours on the dashboard itself are computed from the value, not", 11, False, "333333"),
        ("carried in the file.", 11, False, "333333"),
    ]
    for i, (text, size, bold, color) in enumerate(lines, start=1):
        c = instructions.cell(row=i, column=1, value=text)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    # Instructions is read-only end to end — no cell on it is ever meant to be edited.
    instructions.protection.sheet = True

    ws = wb.create_sheet(pillar)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value=f"{pillar} — {quarter_label}")
    title.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    title.protection = LOCKED

    header_row = 3
    headers = ["Metric", "Sub / Category", quarter_label, "Displayed on Dashboard", "Note to checker"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center", wrap_text=True)
        c.protection = LOCKED

    r = header_row + 1
    current_screen = ""
    for entry in rows:
        if entry[0] == "section":
            current_screen = SCREEN_FOR.get(entry[1], "")
            ws.merge_cells(f"A{r}:E{r}")
            c = ws.cell(row=r, column=1, value=entry[1].upper())
            c.font = Font(name="Calibri", size=10.5, bold=True, color=ACCENT)
            c.fill = PatternFill("solid", fgColor=LIGHT)
            c.protection = LOCKED
            r += 1
            continue
        is_kpi_row = entry[0] == "kpi"
        if is_kpi_row:
            _, kid, name, vals = entry
            ws.cell(row=r, column=1, value=f"{kid} — {name}").font = Font(name="Calibri", size=10, bold=True)
            ws.cell(row=r, column=2, value="YTD Actual")
        elif entry[0] == "record":
            _, record_type, label, category, display, vals, _record_field = entry
            ws.cell(row=r, column=1, value=display).font = Font(name="Calibri", size=10)
            ws.cell(row=r, column=2, value=category or "—")
        else:
            _, metric_key, dim, dim2, display, vals = entry
            ws.cell(row=r, column=1, value=display).font = Font(name="Calibri", size=10)
            ws.cell(row=r, column=2, value=dim2 or "—")
        ws.cell(row=r, column=1).protection = LOCKED
        ws.cell(row=r, column=2).protection = LOCKED

        v = vals[qi]
        cell = ws.cell(row=r, column=3, value=v)
        cell.font = Font(name="Calibri", size=10, color="1B4D3E")
        cell.fill = PatternFill("solid", fgColor="FBFDFC")
        cell.number_format = "#,##0.00" if isinstance(v, float) and v != int(v) else "#,##0"
        cell.alignment = Alignment(horizontal="right")
        cell.border = BORDER
        cell.protection = UNLOCKED  # the one cell per row a preparer is meant to touch

        remark = ws.cell(row=r, column=4, value=current_screen)
        remark.font = Font(name="Calibri", size=8.5, italic=True, color="6B7B76")
        remark.alignment = Alignment(wrap_text=True, vertical="center")
        remark.protection = LOCKED

        # Note-to-checker is a genuine free-text input, but only KPI Scorecard rows actually
        # route through the checker queue / carry that note anywhere — leave it editable there
        # and locked (inert) everywhere else so it can't be mistaken for a place to leave notes
        # that nothing reads.
        note_cell = ws.cell(row=r, column=5)
        note_cell.protection = UNLOCKED if is_kpi_row else LOCKED

        for ci in (1, 2, 4, 5):
            ws.cell(row=r, column=ci).border = BORDER
        r += 1

    widths = {1: 44, 2: 16, 3: 14, 4: 38, 5: 26}
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[header_row].height = 30

    # Enable protection last, once every cell's locked/unlocked state is final. No password —
    # this guards against accidental edits, not a determined one, and a forgotten password would
    # be a support headache for no real security gain here.
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.insertHyperlinks = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False

    wb._sheets.sort(key=lambda s: 0 if s.title == "Instructions" else 1)
    return wb


os.makedirs(OUT_DIR, exist_ok=True)
for qi, (qid, qlabel) in enumerate(zip(QUARTERS, QUARTER_LABELS)):
    for pillar, slug in PILLAR_SLUG.items():
        wb = build_workbook(qi, pillar)
        out_path = os.path.join(OUT_DIR, f"pksof_data_entry_{qid}_{slug}.xlsx")
        wb.save(out_path)
        print("Saved:", out_path)

# ---------------------------------------------------------------------------
# 4. Emit the identical row manifest as a TypeScript lookup table, so the parser
#    that reads this workbook back can never drift out of sync with what was written.
# ---------------------------------------------------------------------------


def esc(s):
    return s.replace("\\", "\\\\").replace("'", "\\'")


def js_entry(sheet, entry, screen):
    if entry[0] == "section":
        return None
    if entry[0] == "kpi":
        _, kid, name, _ = entry
        return f"  {{ sheet: '{sheet}', label: '{esc(f'{kid} — {name}')}', sub: 'YTD Actual', screen: '{esc(screen)}', dest: 'kpi', kpiId: '{kid}' }},"
    if entry[0] == "record":
        _, record_type, label, category, display, _, record_field = entry
        return f"  {{ sheet: '{sheet}', label: '{esc(display)}', sub: '{esc(category or '—')}', screen: '{esc(screen)}', dest: 'record', recordType: '{record_type}', recordLabel: '{esc(label)}', recordField: '{record_field}' }},"
    _, metric_key, dim, dim2, display, _ = entry
    sub = esc(dim2) if dim2 else "—"
    return f"  {{ sheet: '{sheet}', label: '{esc(display)}', sub: '{sub}', screen: '{esc(screen)}', dest: 'metric', metricKey: '{metric_key}', dimension: '{esc(dim)}', dimension2: '{esc(dim2)}' }},"


ts_lines = [
    "// AUTO-GENERATED by scripts/generate_data_entry_template.py — do not hand-edit.",
    "// Mirrors the row manifest baked into the 3-pillar Excel template exactly, so a row's",
    "// (sheet, label, sub) always resolves to the same metricKey/dimension the display layer reads.",
    "// `screen` is the same \"where this renders\" remark the generated workbook's own",
    "// \"Displayed on Dashboard\" column shows (see SCREEN_FOR in the generator script).",
    "export type TemplateFieldDest =",
    "  | { dest: 'kpi'; kpiId: string }",
    "  | { dest: 'metric'; metricKey: string; dimension: string; dimension2: string }",
    "  | { dest: 'record'; recordType: string; recordLabel: string; recordField: 'valueNum' | 'valueNum2' | 'textNote' };",
    "",
    "export const TEMPLATE_FIELDS: ({ sheet: string; label: string; sub: string; screen: string } & TemplateFieldDest)[] = [",
]
for sheet, rows in SHEETS.items():
    current_screen = ""
    for entry in rows:
        if entry[0] == "section":
            current_screen = SCREEN_FOR.get(entry[1], "")
            continue
        line = js_entry(sheet, entry, current_screen)
        if line:
            ts_lines.append(line)
ts_lines.append("];")

with open(TS_MANIFEST_PATH, "w") as f:
    f.write("\n".join(ts_lines) + "\n")
field_count = sum(1 for line in ts_lines if line.startswith("  { sheet:"))
print("Saved:", TS_MANIFEST_PATH, f"({field_count} fields)")
